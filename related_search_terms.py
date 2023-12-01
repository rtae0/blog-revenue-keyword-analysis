import json
import os
import pandas as pd
import matplotlib.pyplot as plt
import requests
import hmac
import base64
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import urllib.request
import time
import hashlib
from suggestqueries import get_google_suggestions
from selenium_crawling import scrape_google_results
from matplotlib import font_manager
from sklearn.linear_model import LinearRegression
from api_credentials import API_KEY, SECRET_KEY, CUSTOMER_ID, CLIENT_ID, CLIENT_SECRET

# Signature 클래스: API 호출 시 필요한 서명을 생성합니다.
class Signature:
    @staticmethod
    def generate(timestamp, method, uri, secret_key):
        message = "{}.{}.{}".format(timestamp, method, uri)
        hash = hmac.new(bytes(secret_key, "utf-8"), bytes(message, "utf-8"), hashlib.sha256)
        return base64.b64encode(hash.digest())

# API 헤더 생성 함수
def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(round(time.time() * 1000))
    signature = Signature.generate(timestamp, method, uri, secret_key)
    return {'Content-Type': 'application/json; charset=UTF-8', 'X-Timestamp': timestamp,
            'X-API-KEY': api_key, 'X-Customer': str(customer_id), 'X-Signature': signature}

# 키워드에 대한 전체 문서 수를 가져오는 함수
def get_total_documents(keyword):
    encText = urllib.parse.quote(keyword)
    url = "https://openapi.naver.com/v1/search/webkr.json?query=" + encText
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id", CLIENT_ID)
    request.add_header("X-Naver-Client-Secret", CLIENT_SECRET)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    try:
        if rescode == 200:
            response_body = response.read()
            text = response_body.decode('utf-8')
            total_docs = json.loads(text)['total']
            return total_docs
        else:
            print("Error Code:", rescode)
            return 0
    except Exception as e:
        print("Error:", str(e))
        return 0

# 키워드의 월별 검색 비율을 가져오는 함수
def get_month_ratios(keyword):
    url = "https://openapi.naver.com/v1/datalab/search"

    startDate = "2016-01-01"  # 시작일 설정
    endDate = "2023-11-01"    # 종료일 설정
    timeUnit = "date"         # 시간 단위 설정

    keywordGroups = []

    keywords = []
    keywords.append(keyword)

    keywordGroups.append({"groupName": keyword, "keywords": keywords})

    body = {
        "startDate": startDate,
        "endDate": endDate,
        "timeUnit": timeUnit,
        "keywordGroups": keywordGroups
    }

    body = json.dumps(body, ensure_ascii=False)

    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id", CLIENT_ID)
    request.add_header("X-Naver-Client-Secret", CLIENT_SECRET)
    request.add_header("Content-Type", "application/json")
    response = urllib.request.urlopen(request, data=body.encode("utf-8"))
    rescode = response.getcode()

    if (rescode == 200):
        response_body = response.read()
        return response_body.decode('utf-8')
    else:
        print("Error Code:" + rescode)
        return 0

# 키워드에 대한 결과를 가져오는 함수
def get_results(hintKeywords):
    BASE_URL = 'https://api.searchad.naver.com'
    uri = '/keywordstool'
    method = 'GET'

    params = {}

    params['hintKeywords'] = hintKeywords
    params['showDetail'] = '1'

    r = requests.get(BASE_URL + uri, params=params,
                     headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))

    return pd.DataFrame(r.json()['keywordList'])

hintKeywords = '파이썬독학'

resultdf = get_results(hintKeywords)
origin_len = len(resultdf)

print(resultdf)
# 열 이름 변경
new_column_names = {
    'relKeyword': 'Keyword',
    'monthlyPcQcCnt': '월간검색수(PC)',
    'monthlyMobileQcCnt': '월간검색수(모바일)',
    'monthlyAvePcClkCnt': '광고클릭수(PC)',
    'monthlyAveMobileClkCnt': '광고클릭수(모바일)',
    'monthlyAvePcCtr': '광고클릭률(PC)',
    'monthlyAveMobileCtr': '광고클릭률(모바일)',
    'plAvgDepth': '노출광고수',
    'compIdx': '경쟁정도'
}

# 문자열 형태의 열을 숫자로 변환
resultdf['monthlyPcQcCnt'] = pd.to_numeric(resultdf['monthlyPcQcCnt'], errors='coerce')
resultdf['monthlyMobileQcCnt'] = pd.to_numeric(resultdf['monthlyMobileQcCnt'], errors='coerce')

resultdf = resultdf[resultdf['relKeyword'].str.contains(hintKeywords)]

# 월간검색수가 20건 이상인 행만 유지
resultdf = resultdf[(resultdf['monthlyPcQcCnt'] + resultdf['monthlyMobileQcCnt']).fillna(0) >= 20]
resultdf = resultdf.rename(columns=new_column_names)
resultdf['광고클릭률'] = resultdf['광고클릭률(모바일)'] + resultdf['광고클릭률(PC)']

resultdf.to_excel('origin.xlsx', index=False)

# 포스팅 수, 연관검색어, 월별 검색 비율, 블로그 링크 및 갯수 수집 시작

total_documents = []
blog_links = []
month_ratios = []
suggest_keywords = []
link_count = []

print("문서수 카운팅 및 연관검색어 파싱 시작", len(resultdf))
for i, keyword in enumerate(resultdf['Keyword'], start=1):
    print(f"{i}/{len(resultdf)}: {keyword}")
    total_docs = get_total_documents(keyword)
    total_documents.append(total_docs)
    suggest_keyword = get_google_suggestions(keyword)
    suggest_keywords.append(suggest_keyword)

resultdf['연관검색어'] = suggest_keywords
resultdf['총문서수'] = total_documents
resultdf['총검색량'] = (resultdf['월간검색수(PC)'] + resultdf['월간검색수(모바일)'])
resultdf['경쟁정도_ratio'] = resultdf['총문서수'] / (resultdf['월간검색수(PC)'] + resultdf['월간검색수(모바일)'])

# 경쟁-검색 균형 지수 계산
resultdf['경쟁-검색 균형 지수'] = (resultdf['월간검색수(PC)'] + resultdf['월간검색수(모바일)']) / (resultdf['경쟁정도_ratio'] + 1)
resultdf = resultdf[(resultdf['경쟁-검색 균형 지수']).fillna(0) >= 1]

resultdf = resultdf.sort_values(by='경쟁-검색 균형 지수', ascending=False)

# 상위 10개 행만 유지
resultdf = resultdf.head(10)

print("월별 검색 ratio 및 블로그글 크롤링 시작")
for i, keyword in enumerate(resultdf['Keyword'], start=1):
    print(f"{i}/{len(resultdf)}: {keyword}")
    month_ratio = get_month_ratios(keyword)
    month_ratios.append(month_ratio)
    blog_count, site_count, blog_link = scrape_google_results(keyword)
    blog_links.append(blog_link)
    link_count.append(blog_count)


resultdf['구글 상위 블로그 상위 갯수'] = link_count
resultdf['구글 상위 블로그'] = blog_links
resultdf['월간 ratio'] = month_ratios


resultdf = resultdf.drop('월간검색수(PC)', axis=1)
resultdf = resultdf.drop('월간검색수(모바일)', axis=1)
resultdf = resultdf.drop('경쟁정도', axis=1)
resultdf = resultdf.drop('광고클릭수(모바일)', axis=1)
resultdf = resultdf.drop('광고클릭수(PC)', axis=1)
resultdf = resultdf.drop('광고클릭률(모바일)', axis=1)
resultdf = resultdf.drop('광고클릭률(PC)', axis=1)


resultdf = resultdf[(resultdf['구글 상위 블로그 상위 갯수']).fillna(0) >= 2]
resultdf = resultdf[(resultdf['연관검색어']).fillna(0) != "[]"]

delete_len = len(resultdf)
print(str(origin_len-delete_len)+"개의 행이 삭제되었습니다")
print(resultdf)


font_path = "./NanumGothic.ttf"
font_manager.fontManager.addfont(font_path)
plt.rcParams['font.family'] = 'NanumGothic'
plt.rcParams['font.sans-serif'] = font_manager.FontProperties(fname=font_path).get_name()



cell_data =[]

for index in range(resultdf.shape[0]):
    cell_data.append(json.loads(resultdf.iloc[index, 10]))

for index, item in enumerate(cell_data):
    results_data = item['results'][0]['data']
    title = item['results'][0]['title']

    # 데이터 준비
    print(results_data)
    periods = [d['period'] for d in item['results'][0]['data']]
    ratios = [d['ratio'] for d in item['results'][0]['data']]
    dates = [datetime.strptime(period, "%Y-%m-%d") for period in periods]
    df = pd.DataFrame({'Date': dates, 'Ratio': ratios})



    # 선형 회귀 모델 훈련
    df['DateInt'] = df['Date'].apply(lambda x: x.toordinal())
    model = LinearRegression()
    model.fit(df[['DateInt']], df['Ratio'])

    # 미래 3개월 예측
    last_date = df['Date'].iloc[-1]
    future_dates = [last_date + timedelta(days=x) for x in range(1, 365)]
    future_date_int = [d.toordinal() for d in future_dates]
    future_df = pd.DataFrame({'DateInt': future_date_int})
    future_predictions = model.predict(future_df)

    # 시각화
    fig, ax = plt.subplots(figsize=(15, 5))

    # 원래 데이터
    colors = ['red', 'green', 'blue', 'cyan', 'magenta', 'yellow', 'black', 'orange']
    years = df['Date'].dt.year.unique()
    for i, year in enumerate(years):
        year_data = df[df['Date'].dt.year == year]
        ax.plot(year_data['Date'], year_data['Ratio'], label=f'{year}년도', color=colors[i % len(colors)])

    # 예측 데이터
    ax.plot(future_dates, future_predictions, label='예측 데이터', linestyle='--', color='gray')

    # 계절별 배경 색상 설정
    for year in range(df['Date'].dt.year.min(), df['Date'].dt.year.max() + 1):
        # 봄: 3월 1일부터 5월 31일 핑크
        ax.axvspan(datetime(year, 3, 1), datetime(year, 5, 31), color='lightpink', alpha=0.3)

        # 여름: 6월 1일부터 8월 31일 초록
        ax.axvspan(datetime(year, 6, 1), datetime(year, 8, 31), color='lightgreen', alpha=0.3)

        # 가을: 9월 1일부터 11월 30일 주황
        ax.axvspan(datetime(year, 9, 1), datetime(year, 11, 30), color='orange', alpha=0.3)

        # 겨울: 12월 1일부터 다음 해 2월 28일(윤년 체크 필요) 파랑
        end_day = 29 if (year + 1) % 4 == 0 and ((year + 1) % 100 != 0 or (year + 1) % 400 == 0) else 28
        ax.axvspan(datetime(year, 12, 1), datetime(year + 1, 2, end_day), color='skyblue', alpha=0.3)

    ax.set_title(title+'에 대한 검색 관심도 추이 및 예측')
    ax.legend(loc='upper left')
    plt.tight_layout()
    plt.savefig(f'{index}.png')
    plt.show()


resultdf['연관검색어'] = resultdf['연관검색어'].apply(lambda x: '\n'.join(x) if isinstance(x, list) else x)
resultdf['구글 상위 블로그'] = resultdf['구글 상위 블로그'].apply(lambda x: '\n'.join(x) if isinstance(x, list) else x)


resultdf = resultdf.drop('월간 ratio', axis=1)
resultdf = resultdf.drop('구글 상위 블로그 상위 갯수', axis=1)

resultdf.to_excel('output.xlsx', index=False)


# 엑셀 워크북 생성
wb = load_workbook(filename='output.xlsx')

# 워크북의 첫 번째 워크시트 가져오기
ws = wb.active

# 두 번째 행부터 시작하여 마지막 행까지 반복
for index in range(1, ws.max_row):
    img_path = f"{index-1}.png"

    # 이미지 객체 생성 및 리스트에 추가
    img = Image(img_path)
    # 이미지를 원하는 셀 위치에 삽입
    ws.add_image(img, f"J{index + 1}")
    ws.row_dimensions[index + 1].height = 380

# 엑셀 셀 크기 조정
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 11
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 13
ws.column_dimensions['G'].width = 13
ws.column_dimensions['H'].width = 13
ws.column_dimensions['I'].width = 50
ws.column_dimensions['J'].width = 190
# 엑셀 파일 저장

wb.save('result.xlsx')